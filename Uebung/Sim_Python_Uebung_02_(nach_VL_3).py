#
#               WiSe 2324 - Simulation technischer Systeme mit Python
#                               Übung 02
#
#_______________________________________________________________________
import numpy as np
import scipy as sc
import matplotlib.pyplot  as plt
from matplotlib.ticker import MultipleLocator
#-----------------------
import os
import csv
import sys
import serial
from datetime import datetime
import _thread as thread
import time
import random 
import openpyxl as opx 

from sympy import true
#-----------------------
os.system('clear')
os.system('cls')
print("\n\n\n")
#======================================== Teil 00 =================================
print(f'\n\n=====================\n||   Aufgabe 0.1   ||\n=====================\n')
#==================================================================================
alex={'Integer':[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10], 'Sprachen':['C++','Python','Matlab'], 'Erwartungen':'Phyton Faehigkeiten verbessern'}
print('\n',type(alex),'\n\n',alex)

#======================================== Teil 01 =================================
print(f'\n\n=====================\n||   Aufgabe 1.1   ||\n=====================\n')
#==================================================================================
a=[[1, 2, 3, 4, 5, 6, 7, 8, 9]]
for i in range(2,10):
    temp=np.array(a[0])*i
    a.append(temp.tolist())
for j in range(0,len(a)):
 print(a[j])

#======================================== Teil 01 =================================
print(f'\n\n=====================\n||   Aufgabe 2.1   ||\n=====================\n')
#==================================================================================
string="56,756,756,7,452,34534,7,65,845,35,345,37,56,3,3,523,1,235,135,57,857,845,745,632,56,58,568,5,1"
print (string)
string=string.replace(',','.')
print (string)

#======================================== Teil 01 =================================
print(f'\n\n=====================\n||   Aufgabe 3.1   ||\n=====================\n')
#==================================================================================
n=[]
temp=1.0
print(n)
while temp<42.1:
   n.append(round(temp,1))
   temp+=0.1
print(n)

#======================================== Teil 01 =================================
print(f'\n\n=====================\n||   Aufgabe 4.1   ||\n=====================\n')
#==================================================================================
liste = [] 
for i in range(100): 
    liste.append(random.randint(1,1000)) 
def deleteEven(liste1): 
    ausgabeListe = [] 
    for i in range(len(liste1)): 
        if liste1[i]%2!=0: 
            ausgabeListe.append(liste1[i]) 
    return ausgabeListe 
test = deleteEven(liste)
print(liste)
print()
print(test)
print()

# meine lösung
def loeschen(lst):
   i=0
   print(lst)
   print()
   while i<len(lst):
      if lst[i]%2==0:
         del lst[i]
      else: i+=1
   print(lst)
loeschen(liste)

#======================================== Teil 01 =================================
print(f'\n\n=====================\n||   Aufgabe 5.1   ||\n=====================\n')
#==================================================================================
liste = [] 
for i in range(100): 
    liste.append(random.randint(1,10)) 

def entfernen(lst,zahl):
   j=0
   for i in range(lst.count(zahl)):
      lst.remove(zahl)
      j+=1
   print(j,'Mal wurde',zahl,'entfernt')
   return lst
print(liste)
entfernen(liste,9)
print(liste)

#======================================== Teil 01 =================================
print(f'\n\n=====================\n||   Aufgabe 6.1   ||\n=====================\n')
#==================================================================================
text='Das stimmt, Python ist wirklich eine tolle Programmiersprache!'
print(len(text))
def ver(str):
   data=str
   l=len(str)
   versch=list(range(0,l))
   temp=[]
   print(data)
   #print(l,temp,len(temp))
   for i in range(0,l,2):
      versch[i]=str[i]
   for j in range(1,l,2):
      temp.append(data[j])
   m=1
   for n in range(1,l,2):
      versch[n]=temp[-m]
      m+=1
   aus=versch[0]
   for i in range(1,l):
      aus+=versch[i]
   #print(temp,len(temp))
   print('\n',versch)
   print('\n',aus)
   return versch
print('====================')
def ent(data):
   l=len(data)
   string=[]
   for i in range(1,l,2):
      string.append(data[i])
   #print(string)
   m=1
   for n in range(1,l,2):
      data[n]=string[-m]
      m+=1
   print('\n',data)
   entsch=data[0]
   for i in range(1,l):
      entsch+=data[i]
   print('\n',entsch)

a=ver(text)
ent(a)

#======================================== Teil 01 =================================
print(f'\n\n=====================\n||   Aufgabe 7.1   ||\n=====================\n')
#==================================================================================
def zusammenfuegen(array):
   ausgabe=str(array[0])
   for i in range(1,len(array)):
      ausgabe =str(ausgabe) + '+' + str(array[i])
   print('\n',ausgabe,'\n\n')
   #return ausgabe
liste=['hallo','welt',',','heute','ist','ein','schoener','Tag']
zusammenfuegen(liste)

#======================================== Teil 01 =================================
print(f'\n\n=====================\n||   Aufgabe 8.1   ||\n=====================\n')
#==================================================================================
liste = [] 
for i in range(100): 
   liste.append(random.randint(0,10))
tupel=tuple(liste)

zahl1 = input('Welche Zahl soll gezählt werden? ') 
anzahl = tupel.count(int(zahl1)) 
print("Die Zahl %i tritt insgesamt %i mal auf." % (int(zahl1),anzahl)) 
print(f"Die Zahl {zahl1} tritt insgesamt {anzahl} mal auf.")

#======================================== Teil 01 =================================
print(f'\n\n=====================\n||   Aufgabe 9.1   ||\n=====================\n')
#==================================================================================
liste = []
weiter=True
while weiter:
 liste.append(random.randint(0,6))
 
 print(f'Die Zufallszahl ist {liste[-1]}. Moechten sie weiter wuerfeln?')
 weiter=input('Antworten sie mit "Ja" oder "Nein".\nAntwort: ')
 print("\n")
 if weiter=="Ja": weiter=True
 else: weiter=False

#======================================== Teil 01 =================================
print(f'\n\n=====================\n||   Aufgabe 10.1   ||\n=====================\n')
print(f'!!!    UMBEDINGT NOCHMAL ANSCHAUEN      !!!')
#==================================================================================
import openpyxl as opx 
fileName = r'C:\Users\alex-\OneDrive\Uni und Verbindung\Uni und aktuelles Semester\Aktuelles Semester\WiSe 2324 - Simulation technischer Systeme mit Python\Übung\Übung 2\MeineDaten.xlsx' 
wb = opx.load_workbook(filename=fileName) 
ws = wb['Tabelle1']

#check how long the dataset is 
laufVar = 0 
field = 'A' 
data = '' 
while data != None: 
   laufVar += 1 
   data = ws['A' + str(laufVar)].value 

# Write data to a Dictionary 
meineDaten = {ws['A1'].value: [], ws['B1'].value: [], ws['C1'].value: []} 
for i in range(2,laufVar-1): 
   meineDaten[ws['A1'].value].append(float(ws['A'+str(i)].value)) 
   meineDaten[ws['B1'].value].append(float(ws['B'+str(i)].value)) 
   meineDaten[ws['C1'].value].append(float(ws['C'+str(i)].value))

print(meineDaten)
print('\n')
print(meineDaten[ws['A1'].value])
print('\n')
print(meineDaten[ws['B1'].value])
print('\n')
print(meineDaten[ws['C1'].value])

#======================================== Teil 01 =================================
print(f'\n\n=====================\n||   Aufgabe 11.1   ||\n=====================\n')
#==================================================================================
MyExelFile='ue2_sinus_quadrat_funktion.xlsx'
wb=opx.Workbook()

MySheet=wb.active
MySheet.title='Uebung 2 - Sinus^2 Funktion'
MySheet['A1']='Zeitvektor t'
MySheet['B1']='Funktionswerte von f(x)=sin(t)^2'

t=[0]
f=[0]
for i in range (0,100):
   t.append(t[i]+0.1)
   f.append(np.sin(t[-1])**2)
print(t,'\n',len(t))
print(f,'\n',len(f))

for num,(time,val) in enumerate(zip(t,f)):
   MySheet['A'+str(num+2)]=t[num]
   MySheet['B'+str(num+2)]=f[num]

fileName=r'C:\Users\alex-\OneDrive\Uni und Verbindung\Uni und aktuelles Semester\Aktuelles Semester\WiSe 2324 - Simulation technischer Systeme mit Python\Eigene Programme\Uebung\ue2_sinus_quadrat_funktion.xlsx'
wb.save(fileName)

#======================================== Teil 01 =================================
print(f'\n\n=====================\n||   Aufgabe 13.1   ||\n=====================\n')
#==================================================================================
X=np.linspace(0,10,51)
def Polynom(x,a=1.,b=1.,c=0.):
   fx=[]
   for i in x:
      fx.append(a*i**2+b*i+c)
   return fx

#======================================== Teil 01 =================================
print(f'\n\n=====================\n||   Aufgabe 13.2   ||\n=====================\n')
#==================================================================================
A=2
B=-10
C=4
fx=Polynom(X,A,B,C)
print(fx)
# postprocessing 
plt.plot(X,fx,label="y") 
#Hier werden die Werte zum Plotten übergeben 
plt.legend() 
plt.grid() 
plt.show()

#======================================== Teil 01 =================================
print(f'\n\n=====================\n||   Aufgabe 13.3   ||\n=====================\n')
#==================================================================================
fx=Polynom(X)
print(fx)
# postprocessing 
plt.plot(X,fx,label="y") 
#Hier werden die Werte zum Plotten übergeben 
plt.legend() 
plt.grid() 
plt.show()

#======================================== Teil 01 =================================
print(f'\n\n=====================\n||   Aufgabe 13.4   ||\n=====================\n')
#==================================================================================
def PolyDiff(x,fx):
   y_diff=[]
   for i in range(1,len(fx)-1):
      diff=(fx[i+1]-fx[i-1])/(2*(x[i+1]-x[i]))
      y_diff.append(diff)
   return y_diff


fx=Polynom(X,A,B,C)
fx_diff=PolyDiff(X,fx)
print(fx)
# postprocessing 
plt.plot(X,fx,label="y")
plt.plot(X[1:-1],fx_diff, label="y'")
#Hier werden die Werte zum Plotten übergeben 
plt.legend() 
plt.grid() 
plt.show()

#======================================== Teil 01 =================================
print(f'\n\n=====================\n||   Aufgabe 14.1   ||\n=====================\n')
#==================================================================================
def someFunc(f1,f2,f3):
   summe = f1+f2+f3
   produkt = f1*f2*f3
   maximal = max([f1,f2,f3])
   minimal = min([f1,f2,f3])
   return summe,produkt,[maximal,minimal]
Sum,Prod,Maxima=someFunc(1,2,3)
print(Sum,Prod,Maxima)

#======================================== Teil 01 =================================
print(f'\n\n=====================\n||   Aufgabe 15.1   ||\n=====================\n')
#==================================================================================
fileName = r'C:\Users\alex-\OneDrive\Uni und Verbindung\Uni und aktuelles Semester\Aktuelles Semester\WiSe 2324 - Simulation technischer Systeme mit Python\Übung\Übung 2\data_export.txt' 
def import_data(path):
   infile=open(path)
   lineStr=[]
   for line in infile:
      lineStr.append(line)
   infile.close()
   print(lineStr)

   z1=lineStr[1].split(' ')
   w1=lineStr[2].split(' ')
   w2=w1[1].replace(',','.')

   zeit=[float(i) for i in z1[1].split('-')]
   wert=[float(i) for i in w2.split('-')]

   data={lineStr[0].strip(): {z1[0]:zeit, w1[0]:wert}}

   print(zeit, wert)
   print('\n',data)
import_data(fileName)

#======================================== Teil 01 =================================
print(f'\n\n=====================\n||   Aufgabe 16.1   ||\n=====================\n')
#==================================================================================
nachricht='SBWKRQ#PDFKW#VSDVV$'
def encryption(text):
   text = text.upper()
   encrypted_message = ''
   for character in text:
      num = ord(character)
      new_num = num + 3
      if new_num > ord('Z'):
         new_num = new_num - 26
      new_character = chr(new_num)
      encrypted_message = encrypted_message + new_character
   # Ausgabe
   print(f"Die verschluesselte Nachricht ist: {encrypted_message}")
   return encrypted_message

def decryption(text):
   print(f"Die verschluesselte Nachricht war: {text}")
   entschluesselt=''
   for zeichen in text:
      num=ord(zeichen)
      new_num=num-3
      if num<=ord('C') and num>=ord('A'):
         new_num=new_num+26 # Muss gemacht werden, da sonderzeichen vor den Buchstaben kommen und daher nur x,y,z die -26 abbekommen haben
      new_zeichen=chr(new_num)
      entschluesselt=entschluesselt+new_zeichen
   print(f'Die entschluesselte Nachricht ist: {entschluesselt}')
   return entschluesselt
Botschaft=decryption(nachricht)
ver=encryption(Botschaft)

#
#========================== Ende =======================================
print("\n\n")
#=======================================================================